from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "Lebron": {
        "points": 61,
        "assists": 19,
        "rebounds": 19
},
    "Kevin": {
        "points": 55,
        "assists": 16,
        "rebounds": 18
    },
    "Steph": {
        "points": 62,
        "assists": 16,
        "rebounds": 11
    },

}

wb = Workbook()
ws = wb.active
ws.title = "Career"

headings = ["Name"] + list(data["Lebron"].keys())
ws.append(headings)



for players in data:
    total = list(data[players].values())
    ws.append([players] + total)

for col in range(2, len(data["Lebron"]) + 2):
    char = get_column_letter(col)
    ws[char + "5"] = f"=SUM({char + '2'}:{char + '4'})/{len(data)}"

ws['A5'].value = "Average"
# char1 = get_row_letter(row)
# ws[char1 + "5"] = "Average"

wb.save("Players.xlsx")
