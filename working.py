from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('tim.xlsx')
ws = wb.active

# for row in range(1, 11):
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)

# ws.merge_cells("A1:D1")
# ws.unmerge_cells("A1:D1")
ws.move_range("C1:D11", rows=2, cols=2)
wb.save('tim.xlsx')